"""POR Auswertung - Hauptskript.

Erzeugt alle Plots und Ergebnistabellen im Output/figures/ Verzeichnis.
Ergebnisse werden in results.json gespeichert für die LaTeX-Datei.
"""
from __future__ import annotations
import json
import os
from pathlib import Path

import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from scipy.optimize import curve_fit
import openpyxl

from utils import (
    CSV_FOLDER,
    XLSX_PATH,
    load_cassy,
    fit_damped,
    noise_std,
    damped_cosine,
    student_t_factor,
)

OUT = Path(__file__).resolve().parent.parent
FIG = OUT / 'figures'
TAB = OUT / 'tables'
FIG.mkdir(exist_ok=True)
TAB.mkdir(exist_ok=True)

# Matplotlib defaults for clean plots.
plt.rcParams.update({
    'font.size': 11,
    'figure.figsize': (6.4, 4.2),
    'axes.grid': True,
    'grid.alpha': 0.3,
    'errorbar.capsize': 2.5,
    'figure.constrained_layout.use': True,
})

DECIMAL_COMMA = lambda s: s.replace('.', ',')

results = {}


# ----------------------------------------------------------------------------
# Aufgabe 8 (Stoppuhr): Eigenfrequenz aus 5 Zeitmessungen.
# ----------------------------------------------------------------------------
def stopwatch_analysis():
    """5 Messungen der Zeit für 10 Schwingungen bei 0,35 A."""
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb['3.1']
    t10 = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] is not None:
            t10.append(float(row[1]))
    t10 = np.array(t10)
    n = len(t10)
    mean = float(np.mean(t10))
    sigma_x = float(np.std(t10, ddof=1))
    # Typ A: Student-t korrigiert.  Bei einer periodischen Größe ist
    # die Reaktionszeit-Streuung bereits in der Stichprobenstreuung der
    # 5 Messungen enthalten; ein separater Typ-B-Beitrag würde
    # doppelt zählen (vgl. ABW, Abschn. 4.2).
    t_fac = student_t_factor(n, '68')
    u_A = t_fac * sigma_x
    u_B_react = None  # nicht verwendet (siehe oben)
    # Typ B: Anzeigeauflösung 0,01 s rechteckverteilt.
    u_B_res = 0.01 / (2 * np.sqrt(3))
    u_t10 = float(np.sqrt(u_A**2 + u_B_res**2))

    T = mean / 10.0
    u_T = u_t10 / 10.0
    f = 1.0 / T
    u_f = u_T / T**2
    omega = 2 * np.pi * f
    u_omega = 2 * np.pi * u_f
    return {
        't10_values': t10.tolist(),
        'n': n,
        't10_mean': mean,
        't10_sigma_x': sigma_x,
        't_factor': t_fac,
        'u_A': u_A,
        'u_B_react': u_B_react,
        'u_B_res': u_B_res,
        'u_t10': u_t10,
        'T': T,
        'u_T': u_T,
        'f': f,
        'u_f': u_f,
        'omega_d': omega,
        'u_omega_d': u_omega,
    }


# ----------------------------------------------------------------------------
# Aufgabe 9 (Augenmessung): Dämpfungskonstante aus Amplituden-Halblog-Fit.
# ----------------------------------------------------------------------------
def eye_amplitude_analysis(T_period: float, u_T_period: float):
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb['3.2']
    rows = list(ws.iter_rows(values_only=True))
    n_list, A1_list, A2_list = [], [], []
    for row in rows[2:]:
        if row[1] is None:
            continue
        n_list.append(int(row[1]))
        A1_list.append(float(row[2]))
        A2_list.append(float(row[5]))
    n_arr = np.array(n_list)
    A1 = np.array(A1_list)
    A2 = np.array(A2_list)
    A_mean = (A1 + A2) / 2.0
    u_A = 0.1  # cm, aus Excel notiert

    # Halblog-Fit: ln(A_n) = ln(A_0) - lambda * T * n
    # Maske auf A > 1.5 cm: Bei kleineren Amplituden zerfällt die
    # Schwingung deutlich schneller als exponentiell (Verhältnisse
    # A_{n+1}/A_n fallen von ~0.77 auf <0.65) -- typisches Zeichen für
    # eine zusätzliche näherungsweise konstante (Coulombsche)
    # Reibungskraft, die das einfache lineare Modell verletzt.
    mask = A_mean > 1.5
    n_fit = n_arr[mask]
    A_fit = A_mean[mask]
    ln_A = np.log(A_fit)
    u_ln_A = u_A / A_fit  # einfache Fehlerfortpflanzung

    # Gewichtete lineare Regression.
    w = 1.0 / u_ln_A**2
    S_w = w.sum()
    S_wx = (w * n_fit).sum()
    S_wy = (w * ln_A).sum()
    S_wxx = (w * n_fit**2).sum()
    S_wxy = (w * n_fit * ln_A).sum()
    D = S_w * S_wxx - S_wx**2
    a0 = (S_wxx * S_wy - S_wx * S_wxy) / D
    a1 = (S_w * S_wxy - S_wx * S_wy) / D
    # Parameter-Unsicherheiten aus inverse Kovarianz.
    u_a0 = np.sqrt(S_wxx / D)
    u_a1 = np.sqrt(S_w / D)
    # Aus a1 = -lambda * T  -> lambda = -a1 / T  (T relativ präzise bekannt).
    lam = -a1 / T_period
    u_lam = np.sqrt((u_a1 / T_period)**2 + (a1 * u_T_period / T_period**2)**2)
    tau = 1.0 / lam
    u_tau = u_lam / lam**2

    # Plot
    fig, ax = plt.subplots(figsize=(6.4, 4.2))
    ax.errorbar(n_arr[mask], A_mean[mask], yerr=u_A,
                fmt='o', color='C0', markersize=5,
                label='Mittel aus 2 Durchgängen (im Fit)')
    n_grid = np.linspace(n_arr.min(), n_arr.max(), 200)
    ax.plot(n_grid, np.exp(a0 + a1 * n_grid), 'r-', lw=1.5,
            label=f'Anpassung: $\\lambda T_d = {abs(a1):.3f} \\pm {u_a1:.3f}$')
    excluded = ~mask & (A_mean > 0)
    ax.errorbar(n_arr[excluded], A_mean[excluded], yerr=u_A,
                fmt='x', color='gray', markersize=7,
                label=r'nicht im Fit ($A<1{,}5$\,cm)')
    ax.set_yscale('log')
    ax.set_xlabel('Schwingungszahl $n$')
    ax.set_ylabel('Amplitude $A_n$ / cm')
    ax.legend()
    fig.savefig(FIG / 'aufgabe9_amplituden.pdf')
    plt.close(fig)
    return {
        'n': n_arr.tolist(),
        'A1': A1.tolist(),
        'A2': A2.tolist(),
        'A_mean': A_mean.tolist(),
        'u_A': u_A,
        'slope': a1,
        'u_slope': u_a1,
        'intercept': a0,
        'u_intercept': u_a0,
        'lambda': lam,
        'u_lambda': u_lam,
        'tau': tau,
        'u_tau': u_tau,
    }


# ----------------------------------------------------------------------------
# Aufgabe 3/8/9 (Cassy bei 0,35 A): Eigenfrequenz und Dämpfung.
# ----------------------------------------------------------------------------
def cassy_main_analysis():
    d = load_cassy('3.3.csv')
    sigma_phi = noise_std(d.t_slow, d.phi)
    fit = fit_damped(d.t_slow, d.phi, sigma_phi)

    # Plot: Rohdaten + Fit
    fig, ax = plt.subplots(figsize=(6.6, 4.4))
    pre_mask = d.t_slow < fit['t_release']
    fit_mask = d.t_slow >= fit['t_release']
    ax.plot(d.t_slow[pre_mask], d.phi[pre_mask], '.', color='gray',
            markersize=2, alpha=0.6, label='vor Loslassen')
    ax.plot(d.t_slow[fit_mask], d.phi[fit_mask], '.', color='C0',
            markersize=3, label='Messung')
    t_curve = np.linspace(fit['t_release'], d.t_slow[-1], 1000)
    ax.plot(t_curve, damped_cosine(t_curve - fit['t_release'], *fit['popt']),
            'r-', lw=1.2, label='Anpassung Gl. (2)')
    # Einhüllende
    env_pos = fit['offset'] + abs(fit['A']) * np.exp(-fit['lambda'] * (t_curve - fit['t_release']))
    env_neg = fit['offset'] - abs(fit['A']) * np.exp(-fit['lambda'] * (t_curve - fit['t_release']))
    ax.plot(t_curve, env_pos, 'r--', lw=0.8, alpha=0.6)
    ax.plot(t_curve, env_neg, 'r--', lw=0.8, alpha=0.6)
    ax.set_xlabel('Zeit $t$ / s')
    ax.set_ylabel(r'Winkel $\varphi$ / w.E.')
    ax.legend(loc='upper right')
    fig.savefig(FIG / 'aufgabe8_cassy_fit.pdf')
    plt.close(fig)

    # Halblog-Vergleichsplot mit Augenamplituden
    fig, ax = plt.subplots(figsize=(6.4, 4.2))
    from scipy.signal import find_peaks
    phi_c = d.phi[fit_mask] - fit['offset']
    t_c = d.t_slow[fit_mask] - fit['t_release']
    peaks_p, _ = find_peaks(phi_c, distance=10)
    peaks_n, _ = find_peaks(-phi_c, distance=10)
    all_peaks = np.sort(np.concatenate([peaks_p, peaks_n]))
    t_peaks = t_c[all_peaks]
    a_peaks = np.abs(phi_c[all_peaks])
    ax.semilogy(t_peaks, a_peaks, 'o', color='C0', markersize=4,
                label='Cassy-Amplitudenmaxima')
    tt = np.linspace(0, t_c[-1], 200)
    ax.semilogy(tt, abs(fit['A']) * np.exp(-fit['lambda'] * tt), 'r-',
                label=f'$|A_0| e^{{-\\lambda t}}$, $\\lambda = {fit["lambda"]:.3f}$\\,s$^{{-1}}$')
    ax.set_xlabel(r'Zeit seit Loslassen $t - t_0$ / s')
    ax.set_ylabel(r'$|\varphi - \varphi_{\mathrm{offset}}|$ / w.E.')
    ax.legend()
    fig.savefig(FIG / 'aufgabe9_cassy_halblog.pdf')
    plt.close(fig)
    f_cassy = fit['omega_d'] / (2 * np.pi)
    u_f_cassy = fit['u_omega_d'] / (2 * np.pi)
    T_cassy = 2 * np.pi / fit['omega_d']
    u_T_cassy = 2 * np.pi * fit['u_omega_d'] / fit['omega_d']**2
    tau = 1.0 / fit['lambda']
    u_tau = fit['u_lambda'] / fit['lambda']**2
    return {
        'A': fit['A'],
        'u_A': fit['u_A'],
        'lambda': fit['lambda'],
        'u_lambda': fit['u_lambda'],
        'omega_d': fit['omega_d'],
        'u_omega_d': fit['u_omega_d'],
        'f': f_cassy,
        'u_f': u_f_cassy,
        'T': T_cassy,
        'u_T': u_T_cassy,
        'phase': fit['phase'],
        'u_phase': fit['u_phase'],
        'offset': fit['offset'],
        'u_offset': fit['u_offset'],
        'tau': tau,
        'u_tau': u_tau,
        't_release': fit['t_release'],
        'sigma_phi_pre': sigma_phi,
    }


# ----------------------------------------------------------------------------
# Aufgabe 6/7: Stromabhängigkeit der Dämpfungskonstanten und Eigenfrequenz.
# ----------------------------------------------------------------------------
def current_dependence_analysis():
    files = sorted(
        [f for f in os.listdir(CSV_FOLDER) if f.startswith('3.5') and f.endswith('.csv')],
        key=lambda x: float(x.split()[1].replace(',', '.')))
    I_list = []
    lam_list, u_lam_list = [], []
    omega_list, u_omega_list = [], []
    A_list = []
    for fn in files:
        d = load_cassy(fn)
        I = float(fn.split()[1].replace(',', '.'))
        sigma = noise_std(d.t_slow, d.phi)
        try:
            res = fit_damped(d.t_slow, d.phi, sigma)
        except Exception as e:
            print(f'Skip {fn}: {e}')
            continue
        I_list.append(I)
        lam_list.append(res['lambda'])
        u_lam_list.append(res['u_lambda'])
        omega_list.append(res['omega_d'])
        u_omega_list.append(res['u_omega_d'])
        A_list.append(abs(res['A']))
    I_arr = np.array(I_list)
    lam_arr = np.array(lam_list)
    u_lam_arr = np.array(u_lam_list)
    omega_arr = np.array(omega_list)
    u_omega_arr = np.array(u_omega_list)
    # Unsicherheit auf I: Multimeter 1% + 0,01 A (rechteckverteilt).
    u_I = np.sqrt((0.01 * I_arr / np.sqrt(3))**2 + (0.01 / np.sqrt(3))**2)

    # Fit lambda(I) = c * I^2
    def quad(I, c):
        return c * I**2

    popt, pcov = curve_fit(quad, I_arr, lam_arr, sigma=u_lam_arr, absolute_sigma=False)
    c_fit = popt[0]
    u_c = float(np.sqrt(pcov[0, 0]))

    # Auch mit Offset, falls die Reibung beiträgt
    def quad_off(I, c, b):
        return c * I**2 + b
    popt2, pcov2 = curve_fit(quad_off, I_arr, lam_arr, sigma=u_lam_arr, absolute_sigma=False)
    c2, b2 = popt2
    u_c2, u_b2 = np.sqrt(np.diag(pcov2))

    fig, ax = plt.subplots(figsize=(6.4, 4.2))
    ax.errorbar(I_arr, lam_arr, xerr=u_I, yerr=u_lam_arr,
                fmt='o', color='C0', label='CASSY-Fitwerte', markersize=5)
    I_grid = np.linspace(0, I_arr.max() * 1.05, 200)
    ax.plot(I_grid, quad(I_grid, c_fit), 'r-',
            label=f'$\\lambda = c\\,I^2$,\n$c = {c_fit:.3f} \\pm {u_c:.3f}$\\,s$^{{-1}}$A$^{{-2}}$')
    ax.plot(I_grid, quad_off(I_grid, c2, b2), 'g--', alpha=0.7,
            label=f'$\\lambda = c\\,I^2 + b$,\n$b = {b2:.3f} \\pm {u_b2:.3f}$\\,s$^{{-1}}$')
    ax.set_xlabel('Spulenstrom $I$ / A')
    ax.set_ylabel(r'Dämpfungskonstante $\lambda$ / s$^{-1}$')
    ax.legend(loc='upper left', framealpha=0.9)
    fig.savefig(FIG / 'aufgabe6_lambda_I.pdf')
    plt.close(fig)

    # Implizites omega_0^2 = omega_d^2 + lambda^2 pro Messung
    # (sollte nach Theorie konstant sein).
    omega0_sq_implied = omega_arr**2 + lam_arr**2

    # Drift-Diagnoseplot: omega_0^2 implizit vs Spulenstrom, mit
    # Anfangsamplitude A_0 (Marker-Groesse) als zweite Dimension.
    fig, ax = plt.subplots(figsize=(6.4, 4.0))
    A_arr = np.array(A_list)
    sizes = 30 + 4 * A_arr  # Marker-Groesse proportional zu A_0
    sc = ax.scatter(I_arr, omega0_sq_implied, s=sizes, c=A_arr,
                    cmap='viridis', edgecolors='k', linewidths=0.5,
                    zorder=3)
    cb = fig.colorbar(sc, ax=ax)
    cb.set_label('Anfangsamplitude $A_0$ / w.E.')
    ax.axhline(np.median(omega0_sq_implied[~np.isclose(I_arr, 1.5)]),
               color='gray', ls='--', alpha=0.6,
               label='Median (ohne $I=1{,}5$\\,A)')
    ax.set_xlabel('Spulenstrom $I$ / A')
    ax.set_ylabel(r'$\omega_d^2 + \lambda^2$ / s$^{-2}$')
    ax.legend()
    fig.savefig(FIG / 'aufgabe7_omega0sq_drift.pdf')
    plt.close(fig)

    # Aufgabe 7: omega_d^2 + lambda^2 = omega_0^2 (constant)
    # Der Datenpunkt bei I=1.5 A wird als Ausreisser markiert: das
    # Fitfenster enthält dort nur ~1.8 Schwingungsperioden, sodass die
    # Anpassung an Gl. (5) numerisch instabil ist (die scipy-Statistik
    # unterschätzt die wahre Unsicherheit deutlich).
    outlier_mask = np.isclose(I_arr, 1.5)
    keep = ~outlier_mask

    x_all = lam_arr**2
    y_all = omega_arr**2
    ux_all = 2 * lam_arr * u_lam_arr
    uy_all = 2 * omega_arr * u_omega_arr

    x = x_all[keep]; y = y_all[keep]
    ux = ux_all[keep]; uy = uy_all[keep]

    # Ungewichtete lineare Regression (Methode der kleinsten Quadrate).
    # Die Statistik-Unsicherheiten aus dem CASSY-Fit sind sehr klein,
    # spiegeln aber nur die rein zufaellige Streuung des cosinus-Modells
    # gegen die Daten wider; systematische Modellbeitraege (z.B. leichte
    # Reibung, Nichtlinearitaet) wirken auf alle Datenpunkte aehnlich
    # und werden besser durch die Streuung der Punkte um die Gerade
    # abgebildet.
    n_pts = len(x)
    Sx = x.sum(); Sy = y.sum()
    Sxx = (x*x).sum(); Sxy = (x*y).sum()
    D = n_pts*Sxx - Sx**2
    intercept = (Sxx*Sy - Sx*Sxy)/D
    slope = (n_pts*Sxy - Sx*Sy)/D
    resid = y - (intercept + slope*x)
    sigma_y_fit = float(np.sqrt((resid**2).sum() / (n_pts - 2)))
    u_int = sigma_y_fit * np.sqrt(Sxx/D)
    u_slope = sigma_y_fit * np.sqrt(n_pts/D)

    fig, ax = plt.subplots(figsize=(6.4, 4.2))
    ax.errorbar(x, y, xerr=ux, yerr=uy, fmt='o', color='C0',
                label=r'Cassy-Fitwerte ($I\leq 1{,}4$\,A)', markersize=5)
    ax.errorbar(x_all[outlier_mask], y_all[outlier_mask],
                xerr=ux_all[outlier_mask], yerr=uy_all[outlier_mask],
                fmt='s', color='gray', markersize=7,
                label=r'$I=1{,}5$\,A: $<2$ Schwingungen im Fit')
    x_grid = np.linspace(0, x_all.max() * 1.05, 200)
    ax.plot(x_grid, intercept + slope * x_grid, 'r-',
            label=f'Anpassung: Steigung ${slope:.2f}\\pm{u_slope:.2f}$\n'
                  f'$\\omega_0^2 = {intercept:.3f}\\pm{u_int:.3f}$\\,s$^{{-2}}$')
    ax.plot(x_grid, intercept - x_grid, 'g--', alpha=0.7,
            label='Theorie: Steigung $-1$')
    ax.set_xlabel(r'$\lambda^2$ / s$^{-2}$')
    ax.set_ylabel(r'$\omega_d^2$ / s$^{-2}$')
    ax.legend(loc='lower left')
    fig.savefig(FIG / 'aufgabe7_omega_lambda.pdf')
    plt.close(fig)
    omega_0 = float(np.sqrt(intercept))
    u_omega_0 = float(u_int / (2 * np.sqrt(intercept)))

    # Zusätzlicher direkter Plot omega_d gegen lambda (Aufgabe 7 wörtlich).
    fig, ax = plt.subplots(figsize=(6.4, 4.2))
    ax.errorbar(lam_arr[keep], omega_arr[keep],
                xerr=u_lam_arr[keep], yerr=u_omega_arr[keep],
                fmt='o', color='C0', markersize=5,
                label=r'Cassy-Fitwerte ($I\leq 1{,}4$\,A)')
    ax.errorbar(lam_arr[outlier_mask], omega_arr[outlier_mask],
                xerr=u_lam_arr[outlier_mask], yerr=u_omega_arr[outlier_mask],
                fmt='s', color='gray', markersize=7,
                label=r'$I=1{,}5$\,A (Ausreißer)')
    lam_grid = np.linspace(0, max(lam_arr) * 1.05, 200)
    omega_th = np.sqrt(np.maximum(omega_0**2 - lam_grid**2, 0))
    ax.plot(lam_grid, omega_th, 'r-',
            label=fr'$\sqrt{{\omega_0^2-\lambda^2}}$, $\omega_0={omega_0:.3f}$\,s$^{{-1}}$')
    ax.set_xlabel(r'Dämpfungskonstante $\lambda$ / s$^{-1}$')
    ax.set_ylabel(r'Eigenkreisfrequenz $\omega_d$ / s$^{-1}$')
    ax.legend()
    fig.savefig(FIG / 'aufgabe7_omega_lambda_direkt.pdf')
    plt.close(fig)
    return {
        'I': I_arr.tolist(),
        'u_I': u_I.tolist(),
        'lambda': lam_arr.tolist(),
        'u_lambda': u_lam_arr.tolist(),
        'omega_d': omega_arr.tolist(),
        'u_omega_d': u_omega_arr.tolist(),
        'A0': A_list,
        'omega0sq_implied': omega0_sq_implied.tolist(),
        'c_quad': c_fit,
        'u_c_quad': u_c,
        'c_off': c2,
        'u_c_off': u_c2,
        'b_off': b2,
        'u_b_off': u_b2,
        'omega_0_from_fit': omega_0,
        'u_omega_0_from_fit': u_omega_0,
        'omega0sq_intercept': intercept,
        'u_omega0sq_intercept': u_int,
        'slope_omega2_lambda2': slope,
        'u_slope_omega2_lambda2': u_slope,
    }


# ----------------------------------------------------------------------------
# Aufgabe 10: Resonanzkurve.
# ----------------------------------------------------------------------------
def resonance_analysis():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb['3.4']
    f_sig_list, A_list = [], []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None or row[1] is None:
            continue
        f_sig_list.append(float(row[0]))
        A_list.append(float(row[1]))
    f_sig = np.array(f_sig_list)
    A_meas = np.array(A_list)
    # Echte Antriebsfrequenz: f = f_sig / 3200
    f_drive = f_sig / 3200.0
    omega = 2 * np.pi * f_drive
    # Unsicherheit der Frequenz: Signalgenerator Auflösung 1 Hz, rechteckverteilt
    u_f_sig = 1.0 / (2 * np.sqrt(3))
    u_omega = 2 * np.pi * u_f_sig / 3200.0
    # Unsicherheit der Amplitude: Ablesen 0,1 cm
    u_A_meas = np.full_like(A_meas, 0.1)

    def amp(omega, M_over_theta, omega_0, lam):
        return M_over_theta / np.sqrt((omega_0**2 - omega**2)**2 + 4 * lam**2 * omega**2)

    # Startparameter
    omega_R_est = omega[np.argmax(A_meas)]
    A_max_est = A_meas.max()
    lam_est = 0.13
    M_over_theta_est = A_max_est * 2 * lam_est * omega_R_est  # näherung
    p0 = [M_over_theta_est, omega_R_est, lam_est]

    popt, pcov = curve_fit(amp, omega, A_meas, p0=p0, sigma=u_A_meas,
                           absolute_sigma=False, maxfev=20000)
    M_theta, omega_0, lam = popt
    perr = np.sqrt(np.diag(pcov))
    u_M, u_om0, u_lam = perr

    # Resonanzfrequenz: omega_R = sqrt(omega_0^2 - 2*lam^2)
    omega_R = float(np.sqrt(omega_0**2 - 2 * lam**2))
    # Fehlerfortpflanzung
    domega_R_domega0 = omega_0 / omega_R
    domega_R_dlam = -2 * lam / omega_R
    u_omega_R = float(np.sqrt((domega_R_domega0 * u_om0)**2 + (domega_R_dlam * u_lam)**2))

    # Maximum Amplitude
    omega_d_fit = float(np.sqrt(omega_0**2 - lam**2))
    A_max = float(M_theta / (2 * lam * omega_d_fit))

    # Halbwertsbreite (HWHM): Wert bei dem A = A_max/sqrt(2)
    # Numerisch
    om_grid = np.linspace(omega.min() * 0.5, omega.max() * 1.5, 100000)
    A_grid = amp(om_grid, *popt)
    target = A_max / np.sqrt(2)
    above = A_grid >= target
    if above.any():
        idx_min = np.argmax(above)
        idx_max = len(above) - 1 - np.argmax(above[::-1])
        om_low = om_grid[idx_min]
        om_high = om_grid[idx_max]
        delta_omega = (om_high - om_low) / 2
    else:
        delta_omega = float('nan')
    # tau aus lambda
    tau = 1.0 / lam
    product = tau * delta_omega

    # Plot
    fig, ax = plt.subplots(figsize=(6.6, 4.4))
    ax.errorbar(omega, A_meas, xerr=u_omega, yerr=u_A_meas,
                fmt='o', color='C0', markersize=4.5, label='Messung')
    om_curve = np.linspace(omega.min() * 0.95, omega.max() * 1.05, 500)
    ax.plot(om_curve, amp(om_curve, *popt), 'r-', lw=1.4,
            label='Anpassung Gl. (3)')
    ax.axvline(omega_R, color='gray', ls=':', alpha=0.7,
               label=f'$\\omega_R = {omega_R:.3f}$\\,s$^{{-1}}$')
    ax.set_xlabel(r'Erregerfrequenz $\omega$ / s$^{-1}$')
    ax.set_ylabel('Amplitude $A(\\omega)$ / cm')
    ax.legend()
    fig.savefig(FIG / 'aufgabe10_resonanz.pdf')
    plt.close(fig)
    return {
        'f_signal': f_sig.tolist(),
        'A': A_meas.tolist(),
        'f_drive': f_drive.tolist(),
        'omega': omega.tolist(),
        'u_omega': float(u_omega),
        'u_A': 0.1,
        'M_over_theta': M_theta,
        'u_M_over_theta': u_M,
        'omega_0': omega_0,
        'u_omega_0': u_om0,
        'lambda': lam,
        'u_lambda': u_lam,
        'omega_R': omega_R,
        'u_omega_R': u_omega_R,
        'A_max': A_max,
        'delta_omega': delta_omega,
        'tau': tau,
        'u_tau': u_lam / lam**2,
        'product_tau_delta_omega': product,
    }


def main():
    print('=== Aufgabe 8: Stoppuhr ===')
    sw = stopwatch_analysis()
    print(f'T = {sw["T"]:.4f} +/- {sw["u_T"]:.4f} s')
    print(f'omega_d = {sw["omega_d"]:.4f} +/- {sw["u_omega_d"]:.4f} 1/s')
    results['stopwatch'] = sw

    print('=== Aufgabe 8/9: Cassy 0.35 A ===')
    cm = cassy_main_analysis()
    print(f'omega_d (Cassy) = {cm["omega_d"]:.4f} +/- {cm["u_omega_d"]:.4f} 1/s')
    print(f'lambda  (Cassy) = {cm["lambda"]:.4f} +/- {cm["u_lambda"]:.4f} 1/s')
    print(f'tau     (Cassy) = {cm["tau"]:.3f} +/- {cm["u_tau"]:.3f} s')
    results['cassy_main'] = cm

    print('=== Aufgabe 9: Augenamplituden ===')
    eye = eye_amplitude_analysis(sw['T'], sw['u_T'])
    print(f'lambda (Auge) = {eye["lambda"]:.4f} +/- {eye["u_lambda"]:.4f} 1/s')
    print(f'tau    (Auge) = {eye["tau"]:.3f} +/- {eye["u_tau"]:.3f} s')
    results['eye'] = eye

    print('=== Aufgabe 6/7: lambda(I) und omega(lambda) ===')
    cd = current_dependence_analysis()
    print(f'c (lambda = c I^2) = {cd["c_quad"]:.4f} +/- {cd["u_c_quad"]:.4f}')
    print(f'omega_0 from omega_d vs lambda intercept = {cd["omega_0_from_fit"]:.4f} +/- {cd["u_omega_0_from_fit"]:.4f}')
    results['current_dep'] = cd

    print('=== Aufgabe 10: Resonanzkurve ===')
    rc = resonance_analysis()
    print(f'omega_0 (Resonanz) = {rc["omega_0"]:.4f} +/- {rc["u_omega_0"]:.4f}')
    print(f'lambda  (Resonanz) = {rc["lambda"]:.4f} +/- {rc["u_lambda"]:.4f}')
    print(f'omega_R = {rc["omega_R"]:.4f} +/- {rc["u_omega_R"]:.4f}')
    print(f'Delta omega = {rc["delta_omega"]:.4f}')
    print(f'tau * Delta omega = {rc["product_tau_delta_omega"]:.3f}')
    results['resonance'] = rc

    # JSON output
    with open(OUT / 'results.json', 'w') as f:
        json.dump(results, f, indent=2, default=str)
    print(f'\nResults written to {OUT / "results.json"}')


if __name__ == '__main__':
    main()

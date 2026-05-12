"""Erzeugt LaTeX-Rohdatentabellen für den Anhang."""
from __future__ import annotations
import json
import os
from pathlib import Path

import numpy as np
import openpyxl

from utils import XLSX_PATH, CSV_FOLDER

OUT = Path(__file__).resolve().parent.parent
TAB = OUT / 'tables'
TAB.mkdir(exist_ok=True)


def write_t10_table():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb['3.1']
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[1] is not None:
            rows.append((int(r[0]), float(r[1])))
    lines = [
        r'\begin{tabular}{c S[table-format=2.2]}',
        r'\toprule',
        r'Durchlauf & {Zeit für 10 Schwingungen $t_{10}$ / \si{\second}} \\',
        r'\midrule',
    ]
    for i, t in rows:
        s = f'{t:.2f}'.replace('.', ',')
        lines.append(f'{i} & {s} \\\\')
    lines.append(r'\bottomrule')
    lines.append(r'\end{tabular}')
    (TAB / 't10.tex').write_text('\n'.join(lines))


def write_amplitudes_table():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb['3.2']
    rows = list(ws.iter_rows(values_only=True))
    data = []
    for row in rows[2:]:
        if row[1] is None:
            continue
        data.append((int(row[1]), float(row[2]), float(row[5])))
    lines = [
        r'\begin{tabular}{c S[table-format=2.1] S[table-format=2.1]}',
        r'\toprule',
        r'Schwingungszahl $n$ & {$A_1$ / \si{\centi\metre}} & {$A_2$ / \si{\centi\metre}} \\',
        r'\midrule',
    ]
    for n, a1, a2 in data:
        s1 = f'{a1:.1f}'.replace('.', ',')
        s2 = f'{a2:.1f}'.replace('.', ',')
        lines.append(f'{n} & {s1} & {s2} \\\\')
    lines.append(r'\bottomrule')
    lines.append(r'\end{tabular}')
    (TAB / 'amplitudes.tex').write_text('\n'.join(lines))


def write_resonance_table():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb['3.4']
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None or row[1] is None:
            continue
        rows.append((float(row[0]), float(row[1])))
    lines = [
        r'\begin{tabular}{S[table-format=4.0] S[table-format=1.2] S[table-format=1.4] S[table-format=1.4]}',
        r'\toprule',
        r'{$f_{\mathrm{Sig}}$ / \si{\hertz}} & {$A$ / \si{\centi\metre}} & {$f_{\mathrm{Antrieb}}$ / \si{\hertz}} & {$\omega$ / \si{\per\second}} \\',
        r'\midrule',
    ]
    for fs, a in rows:
        fdrv = fs / 3200.0
        om = 2 * 3.141592653589793 * fdrv
        sa = f'{a:.2f}'.replace('.', ',')
        sf = f'{fdrv:.4f}'.replace('.', ',')
        som = f'{om:.4f}'.replace('.', ',')
        lines.append(f'{int(fs)} & {sa} & {sf} & {som} \\\\')
    lines.append(r'\bottomrule')
    lines.append(r'\end{tabular}')
    (TAB / 'resonance.tex').write_text('\n'.join(lines))


def write_lambda_I_table():
    with open(OUT / 'results.json') as f:
        r = json.load(f)
    cd = r['current_dep']
    I = cd['I']
    lam = cd['lambda']; ul = cd['u_lambda']
    om = cd['omega_d']; uom = cd['u_omega_d']
    om0sq = cd['omega0sq_implied']

    def fmt_unc(v, u, decimals=4):
        fmt_str = f'{{:.{decimals}f}}'
        return fmt_str.format(v).replace('.', ','), fmt_str.format(u).replace('.', ',')

    lines = [
        r'\begin{tabular}{S[table-format=1.1] c c c}',
        r'\toprule',
        r'{$I$ / \si{\ampere}} & $\lambda$ / \si{\per\second} & $\omega_d$ / \si{\per\second} & $\omega_d^2{+}\lambda^2$ / \si{\per\square\second} \\',
        r'\midrule',
    ]
    for i, l, ul_, o, uo, om0 in zip(I, lam, ul, om, uom, om0sq):
        si = f'{i:.1f}'.replace('.', ',')
        sl_v, sl_u = fmt_unc(l, ul_, decimals=4)
        so_v, so_u = fmt_unc(o, uo, decimals=4)
        # Fehlerfortpflanzung: u(om^2 + lam^2) = sqrt((2*om*u_om)^2 + (2*lam*u_lam)^2)
        u_om0 = (4 * o**2 * uo**2 + 4 * l**2 * ul_**2) ** 0.5
        som0_v, som0_u = fmt_unc(om0, u_om0, decimals=4)
        lines.append(f'{si} & ${sl_v} \\pm {sl_u}$ & ${so_v} \\pm {so_u}$ & ${som0_v} \\pm {som0_u}$ \\\\')
    lines.append(r'\bottomrule')
    lines.append(r'\end{tabular}')
    (TAB / 'lambda_I.tex').write_text('\n'.join(lines))


def write_vergleich_table():
    """Methoden-Vergleichstabelle für die Diskussion."""
    with open(OUT / 'results.json') as f:
        r = json.load(f)
    sw = r['stopwatch']; cm = r['cassy_main']
    eye = r['eye']; cd = r['current_dep']; rc = r['resonance']

    def fmt(v, u):
        """Format with 2 sig figs on uncertainty, value matched."""
        import math
        exp = int(math.floor(math.log10(abs(u))))
        factor = 10 ** (exp - 1)
        u_r = round(abs(u) / factor) * factor
        decimals = max(0, -(exp - 1))
        fmt_str = f'{{:.{decimals}f}}'
        v_s = fmt_str.format(v).replace('.', ',')
        u_s = fmt_str.format(u_r).replace('.', ',')
        return v_s + r' \pm ' + u_s

    DASH = r'\textemdash'
    rows = [
        ('Stoppuhr, 10 Schwingungen', '$\\omega_d$',
         fmt(sw['omega_d'], sw['u_omega_d']), None),
        ('CASSY, freie Schwingung', '$\\omega_d$',
         fmt(cm['omega_d'], cm['u_omega_d']),
         fmt(cm['lambda'], cm['u_lambda'])),
        ('Augenamplituden', None, None,
         fmt(eye['lambda'], eye['u_lambda'])),
        ('Achsenabschnitt $\\omega_d^2{-}\\lambda^2$', '$\\omega_0$',
         fmt(cd['omega_0_from_fit'], cd['u_omega_0_from_fit']), None),
        ('Resonanzkurve', '$\\omega_0$',
         fmt(rc['omega_0'], rc['u_omega_0']),
         fmt(rc['lambda'], rc['u_lambda'])),
    ]
    lines = [
        r'\begin{tabular}{lcll}',
        r'\toprule',
        r'Methode & Größe & {Wert / \si{\per\second}} & {$\lambda$ / \si{\per\second}} \\',
        r'\midrule',
    ]
    for method, sym, value, lam in rows:
        sym_s = sym if sym else DASH
        value_s = f'${value}$' if value else DASH
        lam_s = f'${lam}$' if lam else DASH
        lines.append(f'{method} & {sym_s} & {value_s} & {lam_s} \\\\')
    lines.append(r'\bottomrule')
    lines.append(r'\end{tabular}')
    (TAB / 'vergleich.tex').write_text('\n'.join(lines))


if __name__ == '__main__':
    write_t10_table()
    write_amplitudes_table()
    write_resonance_table()
    write_lambda_I_table()
    write_vergleich_table()
    print('Tables written to', TAB)
